"""
views.py — Full Production Views File
Covers: Auth, Dashboard, Charts, Import, Employees, Leave Management
"""

import pandas as pd
from datetime import date, timedelta, datetime

from django.contrib.auth import authenticate
from django.db.models import Q, Count
from django.core.exceptions import ValidationError
from django.utils import timezone

from rest_framework.decorators import api_view, permission_classes, throttle_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.throttling import AnonRateThrottle
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.exceptions import TokenError

from .models import (
    Employee, Division, User,
    LeaveBalance, LeaveRequest, LeaveAdjustmentLog,
    EmployeeDocument,
)

import secrets
import string
import io
import openpyxl
from django.http import HttpResponse, FileResponse
import os


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def safe_date(value):
    """Parse any date value from Excel safely. Returns a date object (not datetime).
    Use for DateField columns only (dob, doa, expiry dates, etc.).
    For DateTimeField columns use safe_datetime() instead."""
    try:
        if value is None or value == "":
            return None
        if pd.isna(value):
            return None
        d = pd.to_datetime(value, errors="coerce", dayfirst=True)
        return d.date() if not pd.isna(d) else None
    except Exception:
        return None


def safe_datetime(value):
    """Parse any datetime value from Excel safely.
    Returns a timezone-aware datetime (UTC) so Django never emits a
    'received a naive datetime' RuntimeWarning for DateTimeField columns."""
    try:
        if value is None or value == "":
            return None
        if pd.isna(value):
            return None
        d = pd.to_datetime(value, errors="coerce", dayfirst=True)
        if pd.isna(d):
            return None
        # Make timezone-aware in UTC — avoids RuntimeWarning when USE_TZ=True
        return timezone.make_aware(d.to_pydatetime(), timezone.utc)
    except Exception:
        return None


def safe_get(row, key):
    """Get string value from row safely."""
    val = row.get(key)
    try:
        if pd.isna(val) or val is None:
            return ""
    except (TypeError, ValueError):
        pass
    return str(val).strip() if val is not None else ""


def safe_float(val):
    """Parse float, return None if invalid."""
    try:
        f = float(val)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def safe_bool(val):
    """Parse boolean from 0/1/True/False/Yes/No."""
    if val is None or val == "":
        return False
    try:
        if isinstance(val, bool):
            return val
        s = str(val).strip().lower()
        return s in ("1", "true", "yes", "y")
    except Exception:
        return False


def generate_password(length=10):
    """Generate a secure random password with letters, digits, and symbols."""
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    while True:
        pwd = ''.join(secrets.choice(alphabet) for _ in range(length))
        if (any(c.isupper() for c in pwd)
                and any(c.isdigit() for c in pwd)
                and any(c in "!@#$%" for c in pwd)):
            return pwd



# ─────────────────────────────────────────────
# CUSTOM THROTTLE — LOGIN
# ─────────────────────────────────────────────

class LoginRateThrottle(AnonRateThrottle):
    """
    Applies the 'login' rate limit scope defined in settings.py:
        'login': '10/minute'
    Throttles by IP address (REMOTE_ADDR). Returns HTTP 429 when
    the limit is exceeded, with a Retry-After header so clients
    know exactly how long to wait.
    """
    scope = 'login'


# ─────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────

@api_view(['POST'])
@throttle_classes([LoginRateThrottle])
def login_view(request):
    """
    Authenticate and return JWT tokens.

    Rate-limited to 10 attempts per minute per IP via LoginRateThrottle.
    On the 11th attempt within 60 s, DRF returns HTTP 429 with a
    Retry-After header automatically — no extra code needed here.

    Security note: we always return the same generic "Invalid credentials"
    message whether the username doesn't exist OR the password is wrong.
    This prevents username enumeration (an attacker learning which emp_ids
    are valid accounts by observing different error messages).
    """
    username_input = request.data.get("username")
    password = request.data.get("password")

    if not username_input or not password:
        return Response(
            {"status": "error", "message": "Username and password required"},
            status=400,
        )

    # Primary authentication attempt
    user = authenticate(username=username_input, password=password)

    # Secondary attempt: in case the input was an emp_id rather than username
    # (kept for backwards compatibility — both paths produce the same error)
    if user is None:
        try:
            emp_user = User.objects.get(username=username_input)
            user = authenticate(username=emp_user.username, password=password)
        except User.DoesNotExist:
            pass

    if user:
        emp = getattr(user, "employee_profile", None)
        refresh = RefreshToken.for_user(user)

        return Response({
            "status":               "success",
            "username":             user.username,
            "role":                 user.role,
            "must_change_password": user.must_change_password,
            "emp_id":               emp.emp_id if emp else None,
            "access":               str(refresh.access_token),
            "refresh":              str(refresh),
        })

    # Generic message — intentionally does NOT reveal whether the username exists
    return Response(
        {"status": "error", "message": "Invalid credentials"},
        status=401,
    )


# ─────────────────────────────────────────────
# LOGOUT — blacklists the refresh token
# ─────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    """
    Invalidate the supplied refresh token so it cannot be used to obtain
    new access tokens after logout.  The client must also clear its own
    localStorage / cookie store.
    """
    refresh_token = request.data.get('refresh')
    if not refresh_token:
        return Response({'error': 'refresh token is required'}, status=400)

    try:
        token = RefreshToken(refresh_token)
        token.blacklist()          # requires rest_framework_simplejwt.token_blacklist in INSTALLED_APPS
    except TokenError as e:
        return Response({'error': str(e)}, status=400)

    return Response({'message': 'Successfully logged out.'})


# ─────────────────────────────────────────────
# CHANGE PASSWORD
# ─────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user
    new_password = request.data.get("new_password")
    confirm      = request.data.get("confirm_password")

    if not new_password or not confirm:
        return Response({"error": "Both fields required"}, status=400)
    if new_password != confirm:
        return Response({"error": "Passwords do not match"}, status=400)
    if len(new_password) < 8:
        return Response({"error": "Password must be at least 8 characters"}, status=400)

    user.set_password(new_password)
    user.must_change_password = False
    user.save()

    return Response({"message": "Password changed successfully"})


# ─────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_view(request):
    division_name = request.GET.get("division")

    if not division_name:
        return Response({"error": "Division required"}, status=400)

    today        = date.today()
    next_30_days = today + timedelta(days=30)
    next_60_days = today + timedelta(days=60)
    next_90_days = today + timedelta(days=90)

    if division_name == "all":
        employees = Employee.objects.all()
    else:
        employees = Employee.objects.filter(division__name=division_name)

    total    = employees.count()
    active   = employees.filter(is_active=True).count()
    inactive = employees.filter(is_active=False).count()

    active_employees = employees.filter(is_active=True)

    wp_expiring = active_employees.filter(
        wp_expiry__range=(today, next_60_days)
    ).count()

    passport_expiring = active_employees.filter(
        passport_expiry__range=(today, next_90_days)
    ).count()

    incomplete_profiles = employees.filter(
        Q(phone__isnull=True)          | Q(phone="") |
        Q(nationality__isnull=True)    | Q(nationality="") |
        Q(dob__isnull=True)            |
        Q(passport_no__isnull=True)    | Q(passport_no="") |
        Q(work_permit_no__isnull=True) | Q(work_permit_no="") |
        Q(date_joined_company__isnull=True)
    ).count()

    return Response({
        "total_employees":     total,
        "active_employees":    active,
        "inactive_employees":  inactive,
        "wp_expiring":         wp_expiring,
        "passport_expiring":   passport_expiring,
        "incomplete_profiles": incomplete_profiles,
    })


# ─────────────────────────────────────────────
# EMPLOYEE DASHBOARD
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_dashboard_view(request):
    if not hasattr(request.user, 'employee_profile'):
        return Response({"error": "Unauthorized. Employee profile not found."}, status=403)
    
    if request.user.role != "employee":
        return Response({"error": "Forbidden. Only employees can access this dashboard."}, status=403)

    emp = request.user.employee_profile
    today = date.today()

    # Get Leave Balance
    try:
        balance = LeaveBalance.objects.get(employee=emp, year=today.year)
        leave_balance_remaining = balance.annual_remaining + balance.casual_remaining + balance.medical_remaining
    except LeaveBalance.DoesNotExist:
        leave_balance_remaining = 0

    # Get Pending Requests Count
    pending_requests = LeaveRequest.objects.filter(employee=emp, status=LeaveRequest.STATUS_PENDING).count()

    # Get Upcoming Leaves
    upcoming_leaves = LeaveRequest.objects.filter(
        employee=emp, 
        status=LeaveRequest.STATUS_APPROVED, 
        start_date__gte=today
    ).count()

    # Recent leaves (limit 5)
    recent_leaves_qs = LeaveRequest.objects.filter(employee=emp).order_by('-created_at')[:5]
    recent_leaves = [
        {
            "id": lr.id,
            "type": lr.get_leave_type_display(),
            "start_date": lr.start_date,
            "end_date": lr.end_date,
            "total_days": lr.total_days,
            "status": lr.get_status_display()
        }
        for lr in recent_leaves_qs
    ]

    return Response({
        "user": {
            "name": emp.name,
            "emp_id": emp.emp_id,
            "role": emp.designation_aug or emp.designation_ipa or "Employee",
            "division": emp.division.name if emp.division else "N/A"
        },
        "summary": {
            "leave_balance": leave_balance_remaining,
            "pending_requests": pending_requests,
            "upcoming_leaves": upcoming_leaves
        },
        "documents": {
            "passport_expiring": emp.passport_expiring_soon,
            "wp_expiring": emp.wp_expiring_soon
        },
        "recent_leaves": recent_leaves
    })


# ─────────────────────────────────────────────
# CHARTS
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def chart_division_distribution(request):
    division_param = request.GET.get("division")

    if division_param and division_param != "all":
        employees = Employee.objects.filter(division__name=division_param)
    else:
        employees = Employee.objects.all()

    division_data = (
        employees
        .values('division__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    return Response({
        'labels': [item['division__name'] for item in division_data],
        'datasets': [{
            'data': [item['count'] for item in division_data],
            'backgroundColor': [
                '#3B82F6', '#10B981', '#F59E0B', '#EF4444',
                '#8B5CF6', '#06B6D4', '#84CC16', '#F97316',
            ],
            'borderWidth': 1,
        }],
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def chart_monthly_growth(request):
    from django.db.models.functions import TruncMonth

    division_param = request.GET.get("division")

    if division_param and division_param != "all":
        employees = Employee.objects.filter(division__name=division_param)
    else:
        employees = Employee.objects.all()

    # Use timezone.now() (aware datetime) not date.today() (naive date) when filtering
    # a DateTimeField — passing a plain date causes Django to emit a RuntimeWarning.
    now               = timezone.now()
    twelve_months_ago = now - timedelta(days=365)

    monthly_data = (
        employees
        .filter(created_at__gte=twelve_months_ago)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    months  = []
    counts  = []
    current = twelve_months_ago.date().replace(day=1)
    end     = now.date().replace(day=1)

    month_dict = {item['month'].date(): item['count'] for item in monthly_data}

    while current <= end:
        months.append(current.strftime('%b %Y'))
        counts.append(month_dict.get(current, 0))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    return Response({
        'labels': months[-12:],
        'datasets': [{
            'label': 'New Employees',
            'data': counts[-12:],
            'borderColor': '#3B82F6',
            'backgroundColor': 'rgba(59, 130, 246, 0.1)',
            'tension': 0.4,
            'fill': True,
        }],
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def chart_designation_breakdown(request):
    division_param = request.GET.get("division")

    if division_param and division_param != "all":
        employees = Employee.objects.filter(division__name=division_param)
    else:
        employees = Employee.objects.all()

    designation_data = (
        employees
        .exclude(designation_aug__isnull=True)
        .exclude(designation_aug="")
        .values('designation_aug')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    return Response({
        'labels': [item['designation_aug'] for item in designation_data],
        'datasets': [{
            'data': [item['count'] for item in designation_data],
            'backgroundColor': [
                '#3B82F6', '#10B981', '#F59E0B', '#EF4444',
                '#8B5CF6', '#06B6D4', '#84CC16', '#F97316',
                '#EC4899', '#6B7280',
            ],
            'borderWidth': 1,
        }],
    })


# ─────────────────────────────────────────────
# FILE UPLOAD VALIDATION HELPER
# ─────────────────────────────────────────────

# Maximum allowed upload sizes
EXCEL_MAX_BYTES      = 10 * 1024 * 1024   # 10 MB  — Excel imports
ATTACHMENT_MAX_BYTES =  5 * 1024 * 1024   #  5 MB  — Leave attachments

# Allowed Excel magic-byte signatures (first 8 bytes of file)
# xlsx / xlsm → PK zip header (50 4B 03 04)
# xls         → Compound Document header (D0 CF 11 E0 A1 B1 1A E1)
EXCEL_MAGIC = {
    b'\x50\x4b\x03\x04',           # xlsx / xlsm (ZIP-based)
    b'\xd0\xcf\x11\xe0',           # xls  (Compound Document)
}

# Allowed leave-attachment signatures
ATTACHMENT_MAGIC = {
    b'\x25\x50\x44\x46',           # PDF  (%PDF)
    b'\x89\x50\x4e\x47',           # PNG
    b'\xff\xd8\xff',               # JPEG
    b'\x50\x4b\x03\x04',           # DOCX (ZIP-based)
}

EXCEL_EXTENSIONS      = {'.xlsx', '.xls', '.xlsm'}
ATTACHMENT_EXTENSIONS = {'.pdf', '.png', '.jpg', '.jpeg', '.docx'}


def validate_upload(file, allowed_extensions, allowed_magic, max_bytes):
    """
    Three-layer file validation:
      1. Size  — reject files that exceed max_bytes
      2. Extension — reject disallowed file extensions
      3. Magic bytes — read the actual first bytes of the file to confirm
                       the content matches its claimed type.
                       This catches renamed files (e.g. malware.xlsx).

    Returns (True, None) on success.
    Returns (False, "human-readable error message") on failure.
    """
    # Layer 1 — Size check (cheap, do it first)
    if file.size > max_bytes:
        limit_mb = max_bytes // (1024 * 1024)
        return False, f"File too large. Maximum allowed size is {limit_mb} MB."

    # Layer 2 — Extension check (case-insensitive)
    import os
    _, ext = os.path.splitext(file.name.lower())
    if ext not in allowed_extensions:
        return False, (
            f"Invalid file type '{ext}'. "
            f"Allowed: {', '.join(sorted(allowed_extensions))}"
        )

    # Layer 3 — Magic bytes (read first 8 bytes from actual file content)
    header = file.read(8)
    file.seek(0)   # reset so subsequent readers (e.g. Pandas) start at byte 0

    matched = any(header.startswith(magic) for magic in allowed_magic)
    if not matched:
        return False, (
            "File content does not match its extension. "
            "Upload a genuine Excel file."
        )

    return True, None


# ─────────────────────────────────────────────
# EXCEL IMPORT — with auto user creation
# ─────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_excel(request):
    """
    Single Excel upload.
    IS_ACTIVE column: 1 = active (current), 0 = inactive (cancelled).
    If IS_ACTIVE column is missing → defaults to True (active).
    EMP ID is the only required column. All others are optional.

    NEW EMPLOYEES → auto-creates User account with temp password.
    Returns credentials Excel file if any new employees were created.
    Existing employees → updated only, no password change.
    """
    if request.user.role not in ('admin', 'hr'):
        return Response({"error": "Admin or HR access only"}, status=403)

    file = request.FILES.get('file')
    if not file:
        return Response({"error": "No file uploaded"}, status=400)

    # Validate size, extension, and magic bytes before Pandas touches the file
    ok, err = validate_upload(
        file,
        allowed_extensions=EXCEL_EXTENSIONS,
        allowed_magic=EXCEL_MAGIC,
        max_bytes=EXCEL_MAX_BYTES,
    )
    if not ok:
        return Response({"error": err}, status=400)

    try:
        df = pd.read_excel(file, sheet_name=0, dtype=str)
    except Exception as e:
        return Response({"error": f"Excel read failed — {e}"}, status=400)

    # Normalise column names
    df.columns = df.columns.str.strip()
    df = df.where(pd.notnull(df), None)

    created          = 0
    updated          = 0
    skipped          = 0
    errors           = []
    new_credentials  = []   # [{emp_id, name, temp_password}]

    for idx, row in df.iterrows():
        row = row.to_dict()

        # ── EMP ID (required) ──────────────────────────
        emp_id = safe_get(row, "EMP ID")
        if not emp_id:
            skipped += 1
            continue

        # ── IS_ACTIVE ──────────────────────────────────
        raw_active = row.get("IS_ACTIVE")
        if raw_active is None or str(raw_active).strip() == "":
            is_active = True
        else:
            is_active = safe_bool(raw_active)

        # ── DIVISION ───────────────────────────────────
        company  = safe_get(row, "COMPANY").strip().upper() or "UNKNOWN"
        division, _ = Division.objects.get_or_create(name=company)

        emp_name = safe_get(row, "NAME") or f"EMP-{emp_id}"

        # ── Build defaults dict ────────────────────────
        defaults = {
            # Basic
            "name":          emp_name,
            "phone":         safe_get(row, "HP NUMBER") or None,
            "nationality":   safe_get(row, "NATIONALITY") or None,
            "dob":           safe_date(row.get("D.O.B")),
            "qualification": safe_get(row, "QUALIFICATION") or None,
            "division":      division,
            "is_active":     is_active,

            # Job / Salary
            "designation_ipa": safe_get(row, "IPA DESIGNATION") or None,
            "designation_aug": safe_get(row, "Trade") or None,
            "ipa_salary":      safe_float(row.get("IPA SALARY")),
            "per_hr":          safe_float(row.get("PER HR")),
            "salary":          safe_float(row.get("IPA SALARY")),

            # Employment Dates
            "doa":                 safe_date(row.get("DOA")),
            "arrival_date":        safe_date(row.get("ARRIVAL DATE")),
            "date_joined_company": safe_date(row.get("DATE JOINED")),

            # Work Permit / IC
            "work_permit_no": safe_get(row, "IC / WP NO") or None,
            "fin_no":         safe_get(row, "FIN NO") or None,
            "ic_status":      safe_get(row, "IC TYPE") or None,
            "issue_date":     safe_date(row.get("ISSUANCE DATE")),
            "wp_expiry":      safe_date(row.get("S PASS/ WP EXPRIY")),

            # Passport
            "passport_no":     safe_get(row, "PP.NO") or None,
            "passport_expiry": safe_date(row.get("PP EXPIRY")),

            # Certifications / Safety
            "ssic_gt_sn":  safe_get(row, "SSIC GT S/N") or None,
            "ssic_gt_exp": safe_date(row.get("SSIC GT EXP DATE")),
            "ssic_ht_sn":  safe_get(row, "SSIC HT S/N") or None,
            "ssic_ht_exp": safe_date(row.get("SSIC HT EXP DATE")),

            "work_at_height":   safe_bool(row.get("WORK-AT-HEIGHT")),
            "confined_space":   safe_bool(row.get("CONFINED SPACE")),
            "signalman_rigger": safe_bool(row.get("SIGNALMAN & RIGGER COURSE")),
            "welder_no":        safe_get(row, "WELDER NO") or None,
            "lssc_sn":          safe_get(row, "LSSC S/N") or None,

            # Finance / Other
            "bank_account":  safe_get(row, "BANK ACCOUNT NUMBER") or None,
            "accommodation": safe_get(row, "ACCOMODATION") or None,
            "pcp_status":    safe_get(row, "PCP STATUS") or None,
            "remarks":       safe_get(row, "REMARKS") or None,
        }

        try:
            employee, created_flag = Employee.objects.update_or_create(
                emp_id=emp_id,
                defaults=defaults,
            )

            if created_flag:
                created += 1

                # ── Auto-create User for NEW employees only ──
                temp_password = generate_password()

                user_obj, _ = User.objects.get_or_create(
                    username=emp_id,
                    defaults={"role": "employee"}
                )
                user_obj.set_password(temp_password)
                user_obj.must_change_password = True
                user_obj.save()

                # Link user ↔ employee
                employee.user = user_obj
                employee.save(update_fields=["user"])

                new_credentials.append({
                    "emp_id":         emp_id,
                    "name":           emp_name,
                    "temp_password":  temp_password,
                })

            else:
                updated += 1

        except Exception as e:
            errors.append({
                "row":    idx + 2,
                "emp_id": emp_id,
                "error":  str(e),
            })
            skipped += 1
            continue

    # ── If new employees exist → return credentials Excel ──────────
    if new_credentials:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Credentials"

        # Header row
        header_fill  = openpyxl.styles.PatternFill("solid", fgColor="1565C0")
        header_font  = openpyxl.styles.Font(bold=True, color="FFFFFF")
        headers = ["EMP ID", "NAME", "TEMP PASSWORD", "NOTE"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for i, cred in enumerate(new_credentials, 2):
            ws.cell(row=i, column=1, value=cred["emp_id"])
            ws.cell(row=i, column=2, value=cred["name"])
            ws.cell(row=i, column=3, value=cred["temp_password"])
            ws.cell(row=i, column=4, value="Employee must change password on first login")

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 52

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="employee_credentials.xlsx"'
        response["X-Import-Created"]    = str(created)
        response["X-Import-Updated"]    = str(updated)
        response["X-Import-Skipped"]    = str(skipped)
        response["X-Import-Total"]      = str(created + updated + skipped)
        response["Access-Control-Expose-Headers"] = (
            "X-Import-Created, X-Import-Updated, X-Import-Skipped, "
            "X-Import-Total, Content-Disposition"
        )
        return response

    # ── No new employees → normal JSON response ────────────────────
    response_data = {
        "message":    "Import completed",
        "created":    created,
        "updated":    updated,
        "skipped":    skipped,
        "total_rows": created + updated + skipped,
    }
    if errors:
        response_data["errors"] = errors[:20]

    return Response(response_data)


# ─────────────────────────────────────────────
# DIVISIONS
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_divisions(request):
    divisions = Division.objects.filter(is_active=True).values("id", "name")
    return Response(list(divisions))


# ─────────────────────────────────────────────
# EMPLOYEE LIST
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_list(request):
    division     = request.GET.get("division")
    status       = request.GET.get("status")
    search       = request.GET.get("search")
    designation  = request.GET.get("designation")
    nationality  = request.GET.get("nationality")
    expiry_alert = request.GET.get("expiry_alert")
    joined_from  = request.GET.get("joined_from")
    joined_to    = request.GET.get("joined_to")
    incomplete   = request.GET.get("incomplete")

    employees = Employee.objects.select_related("division").all()

    if search:
        employees = employees.filter(
            Q(emp_id__icontains=search) | Q(name__icontains=search)
        )
    if division and division != "all":
        employees = employees.filter(division__name=division)
    if status == "active":
        employees = employees.filter(is_active=True)
    elif status == "inactive":
        employees = employees.filter(is_active=False)
    if designation:
        employees = employees.filter(
            Q(designation_aug__icontains=designation) |
            Q(designation_ipa__icontains=designation)
        )
    if nationality:
        employees = employees.filter(nationality__icontains=nationality)
    if expiry_alert:
        today   = date.today()
        next_60 = today + timedelta(days=60)
        next_90 = today + timedelta(days=90)
        employees = employees.filter(is_active=True)
        if expiry_alert == "wp":
            employees = employees.filter(wp_expiry__range=(today, next_60))
        elif expiry_alert == "passport":
            employees = employees.filter(passport_expiry__range=(today, next_90))
    if joined_from:
        employees = employees.filter(date_joined_company__gte=joined_from)
    if joined_to:
        employees = employees.filter(date_joined_company__lte=joined_to)
    if incomplete == "true":
        employees = employees.filter(
            Q(phone__isnull=True)          | Q(phone="") |
            Q(nationality__isnull=True)    | Q(nationality="") |
            Q(dob__isnull=True)            |
            Q(passport_no__isnull=True)    | Q(passport_no="") |
            Q(work_permit_no__isnull=True) | Q(work_permit_no="") |
            Q(date_joined_company__isnull=True)
        )

    paginator = PageNumberPagination()
    # Cap page_size to prevent full-table dumps via ?page_size=999999
    paginator.page_size = min(int(request.GET.get('page_size', 15)), 100)
    result_page = paginator.paginate_queryset(employees, request)

    is_privileged = request.user.role in ('admin', 'hr')

    data = [
        {
            "emp_id":           e.emp_id,
            "name":             e.name,
            "phone":            e.phone,
            "designation":      e.designation_aug,
            "division":         e.division.name,
            "status":           "Active" if e.is_active else "Inactive",
            # Salary is sensitive — only admin/hr can see it
            **({"salary": e.ipa_salary} if is_privileged else {}),
            "date_joined":      e.date_joined_company,
            "experience_years": e.experience_years,
            "wp_expiry":        e.wp_expiry,
            "passport_expiry":  e.passport_expiry,
        }
        for e in result_page
    ]

    return paginator.get_paginated_response(data)


# ─────────────────────────────────────────────
# EMPLOYEE DETAIL
# ─────────────────────────────────────────────

@api_view(['GET', 'DELETE'])
@permission_classes([IsAuthenticated])
def employee_detail(request, emp_id):
    try:
        e = Employee.objects.select_related("division").get(emp_id=emp_id)
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=404)

    if request.method == 'DELETE':
        # Fix: use role check — not is_staff (a separate, unrelated Django flag)
        if request.user.role != 'admin':
            return Response({"error": "Admin access required"}, status=403)
        emp_info = {
            "emp_id":   e.emp_id,
            "name":     e.name,
            "division": e.division.name if e.division else None
        }
        e.delete()
        return Response({"message": "Employee deleted successfully", "employee": emp_info})

    # GET — enforce access control
    is_privileged = request.user.role in ('admin', 'hr')
    viewer_profile = getattr(request.user, 'employee_profile', None)
    is_own_profile = (
        request.user.role == 'employee'
        and viewer_profile is not None
        and viewer_profile.emp_id == emp_id
    )

    # Regular employees may only view their own record
    if not is_privileged and not is_own_profile:
        return Response({"error": "Permission denied"}, status=403)

    # Base fields — safe for all authenticated users who pass the access check
    response_data = {
        "emp_id":      e.emp_id,
        "name":        e.name,
        "phone":       e.phone,
        "nationality": e.nationality,
        "dob":         e.dob,
        "age":         e.age,
        "division":    e.division.name,
        "status":      "Active" if e.is_active else "Inactive",
        "qualification": e.qualification,

        "designation_ipa":  e.designation_ipa,
        "designation_aug":  e.designation_aug,

        "doa":                 e.doa,
        "arrival_date":        e.arrival_date,
        "date_joined_company": e.date_joined_company,
        "experience_years":    e.experience_years,

        "work_permit_no":   e.work_permit_no,
        "fin_no":           e.fin_no,
        "issue_date":       e.issue_date,
        "ic_status":        e.ic_status,
        "wp_expiry":        e.wp_expiry,
        "wp_expiring_soon": e.wp_expiring_soon,

        "passport_no":            e.passport_no,
        "passport_expiry":        e.passport_expiry,
        "passport_issue_date":    e.passport_issue_date,
        "passport_issue_place":   e.passport_issue_place,
        "passport_expiring_soon": e.passport_expiring_soon,

        "ssic_gt_sn":  e.ssic_gt_sn,
        "ssic_gt_exp": e.ssic_gt_exp,
        "ssic_ht_sn":  e.ssic_ht_sn,
        "ssic_ht_exp": e.ssic_ht_exp,

        "work_at_height":    e.work_at_height,
        "confined_space":    e.confined_space,
        "welder_no":         e.welder_no,
        "lssc_sn":           e.lssc_sn,
        "signalman_rigger":  e.signalman_rigger,
        "firewatchman":      e.firewatchman,
        "gas_meter_carrier": e.gas_meter_carrier,

        "dynamac_pass_sn":  e.dynamac_pass_sn,
        "dynamac_pass_exp": e.dynamac_pass_exp,

        "accommodation":     e.accommodation,
        "pcp_status":        e.pcp_status,
        "security_bond_no":  e.security_bond_no,
        "security_bond_exp": e.security_bond_exp,
        "remarks":           e.remarks,
    }

    # Sensitive financial fields — admin and HR only
    if is_privileged:
        response_data.update({
            "ipa_salary":   e.ipa_salary,
            "per_hr":       e.per_hr,
            "salary":       e.salary,
            "bank_account": e.bank_account,
        })

    return Response(response_data)


# ─────────────────────────────────────────────
# EXPORT EMPLOYEES
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_employees(request):
    # Export contains salary and bank data — admin and HR only
    if request.user.role not in ('admin', 'hr'):
        return Response({"error": "Admin or HR access required"}, status=403)

    division = request.GET.get("division")

    qs = Employee.objects.select_related("division").all()
    if division and division != "all":
        qs = qs.filter(division__name=division)

    data = [
        {
            "EMP ID":            e.emp_id,
            "IS_ACTIVE":         1 if e.is_active else 0,
            "NAME":              e.name,
            "HP NUMBER":         e.phone,
            "NATIONALITY":       e.nationality,
            "D.O.B":             e.dob,
            "COMPANY":           e.division.name,
            "IPA DESIGNATION":   e.designation_ipa,
            "Trade":             e.designation_aug,
            "IPA SALARY":        e.ipa_salary,
            "PER HR":            e.per_hr,
            "DOA":               e.doa,
            "ARRIVAL DATE":      e.arrival_date,
            "DATE JOINED":       e.date_joined_company,
            "EXPERIENCE YEARS":  e.experience_years,
            "IC / WP NO":        e.work_permit_no,
            "FIN NO":            e.fin_no,
            "IC TYPE":           e.ic_status,
            "ISSUANCE DATE":     e.issue_date,
            "S PASS/ WP EXPRIY": e.wp_expiry,
            "PP.NO":             e.passport_no,
            "PP EXPIRY":         e.passport_expiry,
            "SSIC GT S/N":       e.ssic_gt_sn,
            "SSIC GT EXP DATE":  e.ssic_gt_exp,
            "SSIC HT S/N":       e.ssic_ht_sn,
            "SSIC HT EXP DATE":  e.ssic_ht_exp,
            "WORK-AT-HEIGHT":    1 if e.work_at_height else 0,
            "CONFINED SPACE":    1 if e.confined_space else 0,
            "WELDER NO":         e.welder_no,
            "LSSC S/N":          e.lssc_sn,
            "SIGNALMAN & RIGGER COURSE": 1 if e.signalman_rigger else 0,
            "BANK ACCOUNT NUMBER": e.bank_account,
            "ACCOMODATION":      e.accommodation,
            "PCP STATUS":        e.pcp_status,
            "REMARKS":           e.remarks,
        }
        for e in qs
    ]

    return Response(data)


# ─────────────────────────────────────────────
# UPDATE EMPLOYEE
# ─────────────────────────────────────────────

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_employee(request, emp_id):
    # Fix IDOR: only admin or HR may update any employee record
    if request.user.role not in ('admin', 'hr'):
        return Response({"error": "Admin or HR access required"}, status=403)

    try:
        emp = Employee.objects.get(emp_id=emp_id)
    except Employee.DoesNotExist:
        return Response({"error": f"Employee not found: {emp_id}"}, status=404)

    updated_fields = []

    def update(field, key=None):
        k = key or field
        if k in request.data:
            val = request.data[k] or None
            setattr(emp, field, val)
            updated_fields.append(field)

    for f in [
        'phone', 'nationality', 'accommodation', 'qualification', 'remarks',
        'pcp_status', 'bank_account', 'work_permit_no', 'fin_no', 'passport_no',
        'passport_issue_place', 'welder_no', 'lssc_sn', 'dynamac_pass_sn',
        'security_bond_no', 'designation_ipa', 'designation_aug', 'ic_status',
    ]:
        update(f)

    if 'name' in request.data and request.data['name']:
        emp.name = request.data['name']
        updated_fields.append('name')

    for src_key, field_names in [
        ('salary', ['ipa_salary', 'salary']),
        ('per_hr', ['per_hr']),
    ]:
        if src_key in request.data and request.data[src_key]:
            try:
                val = float(request.data[src_key])
            except (ValueError, TypeError):
                return Response({"error": f"Invalid {src_key} format"}, status=400)
            for fn in field_names:
                setattr(emp, fn, val)
                updated_fields.append(fn)

    date_fields = [
        'dob', 'issue_date', 'wp_expiry', 'passport_expiry', 'passport_issue_date',
        'doa', 'arrival_date', 'date_joined_company',
        'ssic_gt_exp', 'ssic_ht_exp', 'dynamac_pass_exp', 'security_bond_exp',
    ]
    for df_field in date_fields:
        if df_field in request.data:
            raw = request.data[df_field]
            if raw:
                try:
                    setattr(emp, df_field, datetime.strptime(raw, '%Y-%m-%d').date())
                    updated_fields.append(df_field)
                except ValueError:
                    return Response(
                        {"error": f"Invalid date for {df_field}. Use YYYY-MM-DD"},
                        status=400,
                    )
            else:
                setattr(emp, df_field, None)
                updated_fields.append(df_field)

    for bf in ['work_at_height', 'confined_space', 'signalman_rigger',
               'firewatchman', 'gas_meter_carrier']:
        if bf in request.data:
            setattr(emp, bf, bool(request.data[bf]))
            updated_fields.append(bf)

    if not updated_fields:
        return Response({"error": "No valid fields to update"}, status=400)

    emp.save()
    return Response({
        "message":          "Employee updated successfully",
        "updated_fields":   updated_fields,
        "experience_years": emp.experience_years,
    })


# ─────────────────────────────────────────────
# LEAVE — BALANCE
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leave_balance(request, emp_id):
    try:
        emp = Employee.objects.get(emp_id=emp_id)
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=404)

    if request.user.role == 'employee':
        profile = getattr(request.user, 'employee_profile', None)
        if not profile or profile.emp_id != emp_id:
            return Response({"error": "Permission denied"}, status=403)

    year = int(request.GET.get('year', date.today().year))

    try:
        balance = LeaveBalance.objects.get(employee=emp, year=year)
    except LeaveBalance.DoesNotExist:
        return Response({"error": f"No leave balance found for {year}"}, status=404)

    return Response({
        "emp_id": emp_id,
        "year":   year,
        "medical": {
            "entitled":  balance.medical_entitled,
            "used":      balance.medical_used,
            "remaining": balance.medical_remaining,
        },
        "casual": {
            "entitled":  balance.casual_entitled,
            "used":      balance.casual_used,
            "remaining": balance.casual_remaining,
        },
        "annual": {
            "entitled":  balance.annual_entitled,
            "used":      balance.annual_used,
            "remaining": balance.annual_remaining,
        },
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leave_balance_adjust(request, emp_id):
    if request.user.role not in ('admin', 'hr'):
        return Response({"error": "HR or Admin access required"}, status=403)

    try:
        emp = Employee.objects.get(emp_id=emp_id)
    except Employee.DoesNotExist:
        return Response({"error": "Employee not found"}, status=404)

    year       = int(request.data.get('year', date.today().year))
    leave_type = request.data.get('leave_type')
    action     = request.data.get('action')
    days       = request.data.get('days')
    note       = request.data.get('note', '')

    if leave_type not in ('medical', 'casual', 'annual'):
        return Response({"error": "Invalid leave_type"}, status=400)
    if action not in ('add', 'deduct'):
        return Response({"error": "action must be 'add' or 'deduct'"}, status=400)
    try:
        days = int(days)
        if days <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return Response({"error": "days must be a positive integer"}, status=400)

    balance, _ = LeaveBalance.objects.get_or_create(
        employee=emp, year=year,
        defaults={
            'medical_entitled': 14,
            'casual_entitled':  7,
            'annual_entitled':  14,
        },
    )

    field_map = {
        'medical': 'medical_used',
        'casual':  'casual_used',
        'annual':  'annual_used',
    }
    field = field_map[leave_type]

    if action == 'add':
        new_val = max(0, getattr(balance, field) - days)
        setattr(balance, field, new_val)
        log_action = 'manual_add'
    else:
        if not balance.has_sufficient_balance(leave_type, days):
            return Response(
                {"error": f"Insufficient {leave_type} balance. Remaining: {balance.get_remaining(leave_type)}"},
                status=400,
            )
        setattr(balance, field, getattr(balance, field) + days)
        log_action = 'manual_deduct'

    balance.save(update_fields=[field, 'updated_at'])

    LeaveAdjustmentLog.objects.create(
        employee=emp,
        leave_request=None,
        action=log_action,
        performed_by=request.user,
        note=note or f"Manual {action} of {days} {leave_type} day(s) for {year}.",
    )

    return Response({
        "message":   f"Balance {action}ed successfully",
        "remaining": balance.get_remaining(leave_type),
    })


# ─────────────────────────────────────────────
# LEAVE — REQUESTS
# ─────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def leave_request_list(request):
    if request.method == 'GET':
        if request.user.role == 'employee':
            profile = getattr(request.user, 'employee_profile', None)
            if not profile:
                return Response({"error": "No employee profile linked"}, status=400)
            qs = LeaveRequest.objects.filter(employee=profile)
        else:
            qs = LeaveRequest.objects.select_related('employee', 'reviewed_by').all()

            emp_id     = request.GET.get('emp_id')
            status     = request.GET.get('status')
            leave_type = request.GET.get('leave_type')
            from_date  = request.GET.get('from_date')
            to_date    = request.GET.get('to_date')

            if emp_id:
                qs = qs.filter(employee__emp_id=emp_id)
            if status:
                qs = qs.filter(status=status)
            if leave_type:
                qs = qs.filter(leave_type=leave_type)
            if from_date:
                qs = qs.filter(start_date__gte=from_date)
            if to_date:
                qs = qs.filter(end_date__lte=to_date)

        paginator = PageNumberPagination()
        # Cap to 100 — prevents full-table dumps via ?page_size=999999
        paginator.page_size = min(int(request.GET.get('page_size', 20)), 100)
        result_page = paginator.paginate_queryset(qs, request)

        data = [
            {
                "id":               lr.id,
                "emp_id":           lr.employee.emp_id,
                "emp_name":         lr.employee.name,
                "leave_type":       lr.get_leave_type_display(),
                "start_date":       lr.start_date,
                "end_date":         lr.end_date,
                "total_days":       lr.total_days,
                "status":           lr.get_status_display(),
                "reason":           lr.reason,
                "reviewed_by":      lr.reviewed_by.username if lr.reviewed_by else None,
                "reviewed_at":      lr.reviewed_at,
                "rejection_reason": lr.rejection_reason,
                "created_at":       lr.created_at,
            }
            for lr in result_page
        ]
        return paginator.get_paginated_response(data)

    # POST
    if request.user.role == 'employee':
        profile = getattr(request.user, 'employee_profile', None)
        if not profile:
            return Response({"error": "No employee profile linked"}, status=400)
        employee = profile
    else:
        emp_id = request.data.get('emp_id')
        if not emp_id:
            return Response({"error": "emp_id is required"}, status=400)
        try:
            employee = Employee.objects.get(emp_id=emp_id)
        except Employee.DoesNotExist:
            return Response({"error": "Employee not found"}, status=404)

    leave_type = request.data.get('leave_type')
    start_raw  = request.data.get('start_date')
    end_raw    = request.data.get('end_date')
    reason     = request.data.get('reason', '')

    if not all([leave_type, start_raw, end_raw]):
        return Response({"error": "leave_type, start_date and end_date are required"}, status=400)

    try:
        start_date = datetime.strptime(start_raw, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_raw,   '%Y-%m-%d').date()
    except ValueError:
        return Response({"error": "Dates must be YYYY-MM-DD"}, status=400)

    if leave_type not in dict(LeaveRequest.LEAVE_TYPE_CHOICES):
        return Response({"error": "Invalid leave_type"}, status=400)

    if leave_type != LeaveRequest.LEAVE_UNPAID:
        try:
            balance = LeaveBalance.objects.get(employee=employee, year=start_date.year)
        except LeaveBalance.DoesNotExist:
            return Response(
                {"error": f"No leave balance record for {start_date.year}. Contact HR."},
                status=400,
            )
        total_days = (end_date - start_date).days + 1
        if not balance.has_sufficient_balance(leave_type, total_days):
            return Response(
                {"error": f"Insufficient {leave_type} balance. Remaining: {balance.get_remaining(leave_type)}, Requested: {total_days}"},
                status=400,
            )

    try:
        lr = LeaveRequest(
            employee=employee,
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date,
            reason=reason,
        )
        if 'attachment' in request.FILES:
            attachment = request.FILES['attachment']
            ok, err = validate_upload(
                attachment,
                allowed_extensions=ATTACHMENT_EXTENSIONS,
                allowed_magic=ATTACHMENT_MAGIC,
                max_bytes=ATTACHMENT_MAX_BYTES,
            )
            if not ok:
                return Response({"error": f"Attachment invalid — {err}"}, status=400)
            lr.attachment = attachment
        lr.save()
    except ValidationError as e:
        return Response({"error": str(e)}, status=400)

    return Response(
        {
            "message":    "Leave request submitted",
            "id":         lr.id,
            "total_days": lr.total_days,
            "status":     lr.get_status_display(),
        },
        status=201,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leave_request_detail(request, pk):
    try:
        lr = LeaveRequest.objects.select_related('employee', 'reviewed_by').get(pk=pk)
    except LeaveRequest.DoesNotExist:
        return Response({"error": "Leave request not found"}, status=404)

    if request.user.role == 'employee':
        profile = getattr(request.user, 'employee_profile', None)
        if not profile or profile.id != lr.employee.id:
            return Response({"error": "Permission denied"}, status=403)

    return Response({
        "id":               lr.id,
        "emp_id":           lr.employee.emp_id,
        "emp_name":         lr.employee.name,
        "leave_type":       lr.get_leave_type_display(),
        "start_date":       lr.start_date,
        "end_date":         lr.end_date,
        "total_days":       lr.total_days,
        "reason":           lr.reason,
        "attachment":       lr.attachment.url if lr.attachment else None,
        "status":           lr.get_status_display(),
        "reviewed_by":      lr.reviewed_by.username if lr.reviewed_by else None,
        "reviewed_at":      lr.reviewed_at,
        "rejection_reason": lr.rejection_reason,
        "created_at":       lr.created_at,
        "updated_at":       lr.updated_at,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leave_request_approve(request, pk):
    if request.user.role not in ('admin', 'hr'):
        return Response({"error": "HR or Admin access required"}, status=403)

    try:
        lr = LeaveRequest.objects.select_related('employee').get(pk=pk)
    except LeaveRequest.DoesNotExist:
        return Response({"error": "Leave request not found"}, status=404)

    if lr.leave_type != LeaveRequest.LEAVE_UNPAID:
        try:
            LeaveBalance.objects.get(employee=lr.employee, year=lr.start_date.year)
        except LeaveBalance.DoesNotExist:
            return Response(
                {"error": f"No leave balance record for {lr.start_date.year}. Create one first."},
                status=400,
            )

    try:
        lr.approve(reviewed_by=request.user)
    except ValidationError as e:
        return Response({"error": str(e)}, status=400)

    return Response({"message": "Leave request approved", "id": lr.id})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leave_request_reject(request, pk):
    if request.user.role not in ('admin', 'hr'):
        return Response({"error": "HR or Admin access required"}, status=403)

    try:
        lr = LeaveRequest.objects.get(pk=pk)
    except LeaveRequest.DoesNotExist:
        return Response({"error": "Leave request not found"}, status=404)

    try:
        lr.reject(reviewed_by=request.user, reason=request.data.get('reason', ''))
    except ValidationError as e:
        return Response({"error": str(e)}, status=400)

    return Response({"message": "Leave request rejected", "id": lr.id})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leave_request_cancel(request, pk):
    try:
        lr = LeaveRequest.objects.select_related('employee').get(pk=pk)
    except LeaveRequest.DoesNotExist:
        return Response({"error": "Leave request not found"}, status=404)

    if request.user.role == 'employee':
        profile = getattr(request.user, 'employee_profile', None)
        if not profile or profile.id != lr.employee.id:
            return Response({"error": "Permission denied"}, status=403)

    try:
        lr.cancel(cancelled_by=request.user)
    except ValidationError as e:
        return Response({"error": str(e)}, status=400)

    return Response({"message": "Leave request cancelled", "id": lr.id})


# ─────────────────────────────────────────────
# LEAVE — AUDIT LOG
# ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leave_audit_log(request):
    if request.user.role == 'employee':
        profile = getattr(request.user, 'employee_profile', None)
        if not profile:
            return Response({"error": "No employee profile linked"}, status=400)
        qs = LeaveAdjustmentLog.objects.filter(employee=profile)
    else:
        qs = LeaveAdjustmentLog.objects.select_related(
            'employee', 'leave_request', 'performed_by'
        ).all()

        emp_id = request.GET.get('emp_id')
        action = request.GET.get('action')
        if emp_id:
            qs = qs.filter(employee__emp_id=emp_id)
        if action:
            qs = qs.filter(action=action)

    paginator = PageNumberPagination()
    # Cap to 100 — prevents full-table dumps via ?page_size=999999
    paginator.page_size = min(int(request.GET.get('page_size', 30)), 100)
    result_page = paginator.paginate_queryset(qs, request)

    data = [
        {
            "id":               log.id,
            "emp_id":           log.employee.emp_id,
            "emp_name":         log.employee.name,
            "action":           log.get_action_display(),
            "leave_request_id": log.leave_request_id,
            "performed_by":     log.performed_by.username if log.performed_by else None,
            "note":             log.note,
            "timestamp":        log.timestamp,
        }
        for log in result_page
    ]

    return paginator.get_paginated_response(data)

# ─────────────────────────────────────────────────────────────────────
# DOCUMENT MANAGEMENT
# ─────────────────────────────────────────────────────────────────────

DOCUMENT_ALLOWED_EXTENSIONS = {'.pdf', '.jpg', '.jpeg', '.png'}
DOCUMENT_MAGIC = {
    b'\x25\x50\x44\x46',           # PDF (%PDF)
    b'\xff\xd8\xff',               # JPEG
    b'\x89\x50\x4e\x47',           # PNG
}
DOCUMENT_MAX_BYTES = 10 * 1024 * 1024   # 10 MB


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def document_list_upload(request, emp_id):
    """
    GET  /api/documents/<emp_id>/  — list all documents for an employee
    POST /api/documents/<emp_id>/  — upload a new document (admin/hr only)

    Employees can GET their own documents only.
    Admin/HR can GET and POST for any employee.
    """
    try:
        emp = Employee.objects.get(emp_id=emp_id)
    except Employee.DoesNotExist:
        return Response({'error': 'Employee not found.'}, status=404)

    is_privileged = request.user.role in ('admin', 'hr')
    is_own        = hasattr(request.user, 'employee_profile') and \
                    request.user.employee_profile.emp_id == emp_id

    if not (is_privileged or is_own):
        return Response({'error': 'Forbidden.'}, status=403)

    # ── GET ──────────────────────────────────────────────────
    if request.method == 'GET':
        docs = EmployeeDocument.objects.filter(employee=emp).select_related('uploaded_by')
        today = date.today()
        data = []
        for d in docs:
            days_left = (d.expiry_date - today).days if d.expiry_date else None
            data.append({
                'id':            d.id,
                'doc_type':      d.doc_type,
                'doc_type_label':d.get_doc_type_display(),
                'label':         d.label or d.get_doc_type_display(),
                'file_name':     os.path.basename(d.file.name),
                'file_path':     d.file.name,   # relative path under MEDIA_ROOT — used for preview URL
                'expiry_date':   d.expiry_date,
                'days_left':     days_left,
                'is_expiring_soon': d.is_expiring_soon,
                'is_expired':    d.is_expired,
                'uploaded_by':   d.uploaded_by.username if d.uploaded_by else '—',
                'uploaded_at':   d.uploaded_at.strftime('%Y-%m-%d'),
                'notes':         d.notes or '',
            })
        return Response(data)

    # ── POST (admin/hr only) ──────────────────────────────────
    if not is_privileged:
        return Response({'error': 'Only admin/HR can upload documents.'}, status=403)

    file      = request.FILES.get('file')
    doc_type  = request.data.get('doc_type', '')
    label     = request.data.get('label', '')
    expiry    = request.data.get('expiry_date') or None
    notes     = request.data.get('notes', '')

    if not file:
        return Response({'error': 'No file uploaded.'}, status=400)
    if doc_type not in ('passport', 'work_permit', 'other'):
        return Response({'error': 'Invalid doc_type.'}, status=400)

    # Reuse the existing 3-layer validate_upload for security
    ok, err = validate_upload(
        file,
        allowed_extensions=DOCUMENT_ALLOWED_EXTENSIONS,
        allowed_magic=DOCUMENT_MAGIC,
        max_bytes=DOCUMENT_MAX_BYTES,
    )
    if not ok:
        return Response({'error': err}, status=400)

    doc = EmployeeDocument.objects.create(
        employee    = emp,
        doc_type    = doc_type,
        label       = label,
        file        = file,
        expiry_date = expiry,
        uploaded_by = request.user,
        notes       = notes,
    )

    return Response({
        'id':          doc.id,
        'doc_type':    doc.doc_type,
        'label':       doc.label or doc.get_doc_type_display(),
        'expiry_date': doc.expiry_date,
        'uploaded_at': doc.uploaded_at.strftime('%Y-%m-%d'),
    }, status=201)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def document_delete(request, pk):
    """DELETE /api/documents/<pk>/delete/  — admin/hr only"""
    if request.user.role not in ('admin', 'hr'):
        return Response({'error': 'Only admin/HR can delete documents.'}, status=403)
    try:
        doc = EmployeeDocument.objects.get(pk=pk)
    except EmployeeDocument.DoesNotExist:
        return Response({'error': 'Document not found.'}, status=404)

    # Remove physical file from disk
    if doc.file and os.path.isfile(doc.file.path):
        os.remove(doc.file.path)
    doc.delete()
    return Response({'message': 'Document deleted.'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def document_download(request, pk):
    """
    GET /api/documents/<pk>/download/
    Serves the file as an attachment.
    Employees can only download their own docs; admin/hr can download any.
    """
    try:
        doc = EmployeeDocument.objects.select_related('employee__user').get(pk=pk)
    except EmployeeDocument.DoesNotExist:
        return Response({'error': 'Document not found.'}, status=404)

    is_privileged = request.user.role in ('admin', 'hr')
    is_own        = hasattr(request.user, 'employee_profile') and \
                    request.user.employee_profile == doc.employee

    if not (is_privileged or is_own):
        return Response({'error': 'Forbidden.'}, status=403)

    if not os.path.isfile(doc.file.path):
        return Response({'error': 'File not found on server.'}, status=404)

    file_name = os.path.basename(doc.file.name)
    response  = FileResponse(open(doc.file.path, 'rb'), as_attachment=True, filename=file_name)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def documents_expiring(request):
    """
    GET /api/documents/expiring/
    Admin/HR: returns all documents expiring within 60 days across all employees.
    """
    if request.user.role not in ('admin', 'hr'):
        return Response({'error': 'Forbidden.'}, status=403)

    from datetime import timedelta
    from django.utils import timezone
    cutoff = date.today() + timedelta(days=60)
    docs = (
        EmployeeDocument.objects
        .filter(expiry_date__isnull=False, expiry_date__lte=cutoff, expiry_date__gte=date.today())
        .select_related('employee', 'employee__division')
        .order_by('expiry_date')
    )
    today = date.today()
    data = [{
        'id':            d.id,
        'emp_id':        d.employee.emp_id,
        'emp_name':      d.employee.name,
        'division':      d.employee.division.name if d.employee.division else '—',
        'doc_type':      d.doc_type,
        'doc_type_label':d.get_doc_type_display(),
        'label':         d.label or d.get_doc_type_display(),
        'expiry_date':   d.expiry_date,
        'days_left':     (d.expiry_date - today).days,
    } for d in docs]
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def document_audit_log(request):
    """
    GET /api/documents/audit/?emp_id=EMP001
    Returns upload/delete activity for documents.
    Proxies from EmployeeDocument — lists uploads as audit events.
    Admin/HR only.
    """
    if request.user.role not in ('admin', 'hr'):
        return Response({'error': 'Forbidden.'}, status=403)

    emp_id = request.GET.get('emp_id')
    qs = EmployeeDocument.objects.select_related('employee', 'uploaded_by')

    if emp_id:
        qs = qs.filter(employee__emp_id=emp_id)

    qs = qs.order_by('-uploaded_at')[:50]

    data = [{
        'id':           d.id,
        'action':       'uploaded',
        'performed_by': d.uploaded_by.username if d.uploaded_by else '—',
        'message':      f"{d.uploaded_by.username if d.uploaded_by else 'Someone'} uploaded {d.get_doc_type_display()} — {d.label or d.file.name.split('/')[-1]}",
        'timestamp':    d.uploaded_at,
    } for d in qs]

    return Response(data)
